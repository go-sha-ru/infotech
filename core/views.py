from django.views.generic import ListView
from drf_spectacular.utils import extend_schema
from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView
from openpyxl import load_workbook

from core.models import Data
from core.serializers import DataSerializer, FileSerializer


class FileUploadView(APIView):
    parser_classes = (MultiPartParser, )

    @extend_schema(
        request=FileSerializer,
        responses=DataSerializer
    )
    def post(self, request):
        up_file = request.FILES['file']
        ret = []
        wb = load_workbook(up_file)
        sheet_names = wb.get_sheet_names()
        if not len(sheet_names):
            return Response(status.HTTP_204_NO_CONTENT)
        sheet = wb.active
        for row in range(2, sheet.max_row + 1):
            ne = sheet.cell(row=row, column=1).value
            address = sheet.cell(row=row, column=2).value
            coordinates = sheet.cell(row=row, column=3).value
            technology = sheet.cell(row=row, column=4).value
            s = sheet.cell(row=row, column=5).value
            [lat, lon] = coordinates.split(', ')
            technology = technology.split(', ') if technology else []
            data, created = Data.objects.get_or_create(ne=ne)
            data.address = address
            data.latitude = lat
            data.longitude = lon
            data.gsm = True if 'gsm' in technology else False
            data.lte = True if 'lte' in technology else False
            data.umts = True if 'umts' in technology else False
            data.status = s
            data.save()
            ret.append(data)
        serializer = DataSerializer(ret, many=True)
        return Response(serializer.data, status.HTTP_201_CREATED)


class DataView(APIView):
    serializer = DataSerializer

    @extend_schema(
        request=DataSerializer,
        responses=DataSerializer
    )
    def get(self, request):
        data = Data.objects.all()
        serializer = self.serializer(data, many=True)
        return Response(serializer.data)


class DataListView(ListView):
    model = Data
    context_object_name = 'data'
    template_name = "base.html"
